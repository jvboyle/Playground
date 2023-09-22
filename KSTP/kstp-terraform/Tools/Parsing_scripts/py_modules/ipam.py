import openpyxl
from py_modules import filehelper

class ParseIPAM():
    def __init__ (self, ipam_output_folder, ipam_xls, ipam_keyword):
        self._err = False
        self._data = ""
        self.ipam_output_folder = ipam_output_folder
        self.ipam_keyword = ipam_keyword
        try:
            self.workbook = openpyxl.load_workbook(ipam_xls, data_only=True)
        except Exception as ex:
            self._err = True
            self._data = ex.__doc__

    # Class properties

    def _get_err(self):
        return self._err

    def _get_data(self):
        return self._data

    ipam_error = property(
        fget = _get_err
    )

    ipam_data = property(
        fget = _get_data
    )

    # Class members

    def parse_worksheet(self, wks_name, log_writer=""):
        # Select the proper function to parse
        if wks_name == "SRES & DRES":
            self._sres_dres(self.workbook["SRES & DRES"], log_writer)
            

    def _sres_dres(self, wks, log_writer=""):
        subnets = []
        ips = []
        # CSV files will be created so, we define their headers first
        SUBNET_HEADER = ["section", "subnet", "description"]
        IP_HEADER = ["section", "ip address", "hostname", "description", "subnet"]
        try:
            myCSV = filehelper.fileHelper()
            # Loop through the subnets' columns
            for col in range(2, wks.max_column):
                if (wks.cell(2, col).value):
                    consumable_service = wks.cell(2, col).value
                    if (self.ipam_keyword in consumable_service.lower()):
                        log_writer.info("   |-- Parsing subnet '{}'".format(consumable_service))
                        name = wks.cell(3, col).value
                        vlan = wks.cell(4, col).value
                        subnets.append({
                            "section": consumable_service,
                            "description": name.replace("VxLAN: ", ""),
                            "subnet": vlan
                        })
                        # For each subnet, get the IPs
                        log_writer.info("   |-- Parsing IPs for '{}', please wait ...".format(consumable_service))
                        for row in range(6, wks.max_row):
                            ip = wks.cell(row, col).value
                            if (ip):
                                # If the cell is strikethrough, ignore
                                if (not wks.cell(row,col).font.strike):
                                    hostname = wks.cell(row, col+1).value 
                                    if (wks.cell(row, col+1).font.strike):
                                        hostname = ""
                                    description = wks.cell(row, col+2).value
                                    if (wks.cell(row, col+2).font.strike):
                                        description = ""
                                    ips.append({
                                        "section": consumable_service,
                                        "ip address": ip,
                                        "hostname": hostname,
                                        "description": description,
                                        "subnet": vlan
                                    })
                        # Save IPs
                        save_as = "{}/{}-IPs.csv".format(self.ipam_output_folder, consumable_service)
                        #save_as = save_as.replace(" ", "-").replace("_", "-")
                        log_writer.info("   |-- Writing '{}'".format(save_as))
                        myCSV.write_csv_file(save_as, IP_HEADER, ips)
                        if (myCSV.file_error):
                            self._err = True
                            self._data = myCSV.file_data
                            log_writer.critical("      |-- Error saving the file")
                            log_writer.error("      |-- Error message: {}".format(myCSV.file_data))
                        else: 
                            log_writer.info("      |-- File successfully saved")
                        ips = []
                    else:
                        if (log_writer):
                            log_writer.info("   |-- Skipping '{}' as it didn't match argument '-k {}'".format(consumable_service, self.ipam_keyword))
            # Save SUBNETS
            save_as = "{}/{}-subnets.csv".format(self.ipam_output_folder, self.ipam_keyword)
            log_writer.info("   |-- Writing '{}'".format(save_as))
            myCSV.write_csv_file(save_as, SUBNET_HEADER, subnets)
            if (myCSV.file_error):
                self._err = True
                self._data = myCSV.file_data
                log_writer.critical("      |-- Error saving the file")
                log_writer.error("      |-- Error message: {}".format(myCSV.file_data))
            
            # If no errors, just inform
            if (not self._err):
                self._data = "Data successfully saved."

        except Exception as ex:
            self._err = True
            self._data = ex.__doc__