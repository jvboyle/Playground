# WOOKBOOK='/home/sdubrouskaya/ISPW/IPAM_SPREDSHEET_MIGRATION/input_spreadsheets/ISPW-REG1-IPAM.xlsx'
# worksheet_name='SRES & DRES'
# reg_ex="r'SR'"
# output="output.json"
# python3.8 parse_tab.py -f "$WOOKBOOK" -s "$reg_ex" -w "$worksheet_name" -o "$output"
# WOOKBOOK='/home/sdubrouskaya/ISPW/IPAM_SPREDSHEET_MIGRATION/input_spreadsheets/ISPW-REG1-IPAM.xlsx'
# worksheet_name='HRES'
# output="output.json"
# python3.8 parse_tab.py -f "$WOOKBOOK"  -w "$worksheet_name" -o "$output"


# Import openyxl module
import openpyxl
import re
import parse
import getopt 
import sys

def get_params(argumentList):
    
    global WOOKBOOK
    global reg_ex
    global worksheet_name
    global file_name
    
    print(argumentList)

    # Options
    options = "hf:s:w:o:"
 
    
    try:
        # Parsing argument
        arguments, values = getopt.getopt(argumentList, options)

        # checking each argument
        for currentArgument, currentValue in arguments:
 
            if currentArgument in ("-h"):
                print ("python3.8 parse_tab.py -f <absolute path to the ipam spreadsheet> -s <regular expresion for section> -w <worksheet mame>")
                print ("DR - regular expression to parse DRES \n" )
                sys.exit(2)
            elif currentArgument in ("-f"):
                 WOOKBOOK = currentValue
            elif currentArgument in ("-s"):
                 reg_ex = currentValue
            elif currentArgument in ("-w"):  
                 worksheet_name = currentValue 
            elif currentArgument in ("-o"):
                 file_name = currentValue           
             
    except getopt.error as err:
        # output error, and return with an error code
        print (str(err))

file_name = ''
get_params(sys.argv[1:])

wookbook = openpyxl.load_workbook(WOOKBOOK)



worksheet = wookbook[worksheet_name]

if worksheet_name == "SRES & DRES":
   info = [] 
   for col in range(2,worksheet.max_column):
      if worksheet.cell(2,col).value and re.match(reg_ex, worksheet.cell(2,col).value):
         if not re.match(r"DRxx",  worksheet.cell(2,col).value):
            info.append(parse.parse_dres_tab(worksheet,col))
       
if worksheet_name == "GSNI CGN Translation": 
   info =  parse.parse_gsni_cgn_translation(worksheet,2)

if worksheet_name == "HRES": 
   info =  parse.parse_hres(worksheet,2)

if worksheet_name == "SL Public Subnets(VL874)": 

   info = [] 
   for col in range(2,worksheet.max_column):
       if worksheet.cell(2,col).value:
          info.append(parse.parse_sl_public_sub(worksheet,col))

if worksheet_name == "Transit & NAT": 

   edge = []
   agg  = []
   for col in range(2,worksheet.max_column):
       if worksheet.cell(2,col).value:
          if re.match(r"Edge Transit Networks", worksheet.cell(2,col).value):            
             edge.append(parse.parse_transit_and_nat(worksheet,col,"edge"))

          if re.match(r"Provider Agg Transit", worksheet.cell(2,col).value):            
             agg.append(parse.parse_transit_and_nat(worksheet,col,"agg"))  

          if re.match(r"IRES Secondary IP", worksheet.cell(2,col).value):            
             agg.append(parse.parse_transit_and_nat(worksheet,col,"ires"))       
          
if worksheet_name == "SL Private Subnets": 
   pass

if worksheet_name == "VDR": 
   pass

if worksheet_name == "VPC": 
   pass

if not file_name:
   file_name = reg_ex + '.json'
parse.write_to_file(file_name,info)   